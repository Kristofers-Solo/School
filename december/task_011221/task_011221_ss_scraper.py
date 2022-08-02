# Author - Kristiāns Francis Cagulis
# Date - 07.12.2021
# Title - Patstāvīgais darbs "SS.com scraping"

from bs4 import BeautifulSoup
import requests
import pandas as pd
from PIL import Image
from io import BytesIO
from openpyxl.styles import Font, Alignment
import openpyxl

HEADERS = {
    "User-Agent": 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/96.0.4664.97 Safari/537.36 Vivaldi/4.1.2369.21'}


class SS:
    def __init__(self, url):
        self.url = url

    def _get_page_amount(self):
        page = requests.get(self.url, headers=HEADERS)
        soup = BeautifulSoup(page.content, 'html.parser')

        last_url = soup.find(class_='td2').findChild('a')['href']
        page_amount = last_url[last_url.find(
            "page") + 4:last_url.find(".html")]
        print(f"Page amount = {page_amount}")

        return int(page_amount)

    def get_data(self):
        items = []
        images = []
        item_no = 1
        for page_number in range(1, self._get_page_amount() + 1):
            url = self.url + f"/page{page_number}.html"

            page = requests.get(url, headers=HEADERS)
            soup = BeautifulSoup(page.content, 'html.parser')

            # item ids
            ids = [tag['id']
                   for tag in soup.select('tr[id]')]  # creates list with ids
            # removes "tr_bnr" elements from list
            ids = [x for x in ids if "tr_bnr" not in x]
            ids.remove("head_line")  # removes first "head_line" id
            print(f"Page {page_number}")

            # getting item data
            for el in soup.find_all(id=ids):
                print(f"Item {item_no}")
                item_no += 1

                # image
                image_url = el.find(class_='msga2').find_next_sibling().findChild(
                    'a').findChild('img')['src']  # gets image url
                response = requests.get(image_url)
                img = Image.open(BytesIO(response.content))
                images.append(img)
                print(img)
                for elem in el.find_all(class_='msga2-o pp6'):
                    items.append(elem.get_text())

                # adverts url
                item_url = el.findChild(class_='msg2').findChild(
                    'div').findChild('a')['href']  # gets url
                item_url = "https://www.ss.com" + item_url
                item_page = requests.get(item_url, headers=HEADERS)
                item_soup = BeautifulSoup(item_page.content, 'html.parser')

                # adverts full text
                item_text = item_soup.find(
                    id='msg_div_msg').get_text()  # gets full text
                # removes text last part (table)
                item_text = item_text[:item_text.find("Pilsēta:")]
                items.append(item_text)

                # adverts publication date
                # gets all 'msg_footer' class'
                item_date = item_soup.find_all('td', class_='msg_footer')
                item_date = item_date[2].get_text()  # extracts 3rd element
                items.append(item_date[8:18])  # crops date

        chunk_size = 8
        # combines each 'chunk_size' elements into array
        chunked_items_list = [items[i:i + chunk_size]
                              for i in range(0, len(items), chunk_size)]
        columns = ["Atrašanās vieta", "Istabu skaits", "Kvadratūra", "Stāvs",
                   "Sērija", "Cena", "Pilns sludinājuma teksts", "Izvietošanas datums"]
        df = pd.DataFrame(chunked_items_list, columns=columns)
        df.to_excel(excel_writer='output.xlsx', index=False)

        wb = openpyxl.load_workbook("output.xlsx")
        ws = wb.worksheets[0]
        sheet = wb.active

        # 'I1' cell setup
        ws['I1'] = "Attēli"
        ws['I1'].font = Font(bold=True)
        ws["I1"].alignment = Alignment(horizontal='center', vertical='top')

        # sets cell width
        sheet.column_dimensions['A'].width = 20
        sheet.column_dimensions['G'].width = 50
        sheet.column_dimensions['H'].width = 20
        sheet.column_dimensions['I'].width = 13

        for i in range(len(images)):
            sheet.row_dimensions[i + 2].height = 51  # sets cell height
            # enables word wrap
            ws[f'G{i + 2}'].alignment = Alignment(wrap_text=True)

            img = openpyxl.drawing.image.Image(images[i])
            ws.add_image(img, f"I{i + 2}")  # adds images
        wb.save("output.xlsx")
        print("Done")


flats = SS("https://www.ss.com/lv/real-estate/flats/riga/all/sell/")
flats2 = SS("https://www.ss.com/lv/real-estate/flats/riga-region/all/sell/")


def main():
    flats.get_data()


if __name__ == '__main__':
    main()
